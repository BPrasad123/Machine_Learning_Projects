#Import necessary Libraries
import pandas as pd
import os
import datetime
from math import ceil




# Set the working directory and import required datasets
os.chdir(r'path to the working directory')
df = pd.read_csv('file_name.csv')




# Rename columns if necessary
cols = list(df.columns)
df = df.rename(columns = {'Variable_A\n' : 'Variable_A',    # if column name has a new line character that is not easily visible
                          'Variable_B\n' : 'Variable_B',
                          'Variable_C\n' : 'Variable_C',
                          'Date_1\n' : 'Date_1',
                          'Date_2\n' : 'Date_2',
                          'Actual_col_name\n' : 'Variable_D',
                          'Actual_col_name_1\n' : 'Variable_E',
                          'Actual_col_name_2' : 'Variable_F',  # renaming to a different name
                          'Actual_col_name_3' : 'Variable_G',
                          'Variable_H\n' : 'Variable_H'
                          })



    
# KPI - Variable_I and Variable_B share


df_temp = df[df['Variable_A'].notnull()]
df_temp['Variable_B'] = df_temp['Variable_B'].fillna(value = 'Others')
df_sys = df_temp.groupby(['Variable_B'])['Variable_A'].nunique()
df_sys = df_sys.reset_index(drop = False)
df_sys.columns = ['Variable_B',	'Variable_I']
df_sys['Variable_B'] = df_sys['Variable_I'].apply(lambda x: "%.1f" % ((x / df_sys['Variable_I'].sum()) * 100))
df_sys = df_sys.sort_values(by=['Variable_I'], ascending=False)
unique_asset_count = df_temp['Variable_A'].nunique()







# KPI - Variable_I and share against Product Variable_C


df_temp = df[df['Variable_A'].notnull()]
df_temp['Variable_J'].nunique()
df_temp['Variable_J'] = df_temp['Variable_J'].fillna(value = 'Others')
df_product = df_temp.groupby(['Variable_J'])['Variable_A'].nunique()
df_product = df_product.reset_index(drop = False)
df_product.columns = [r'Variable_C share',	'Variable_I']
df_product['Variable_C'] = df_sys['Variable_I'].apply(lambda x: "%.1f" % ((x / df_sys['Variable_I'].sum()) * 100))
df_product = df_product.sort_values(by=['Variable_I'], ascending=False)
unique_product_Variable_C_count = df_temp['Variable_J'].nunique()











# KPI - products in warranty; finding how many moving out in next 180 days


df_temp = df[df['Variable_A'].notnull() & (df['Variable_C'] == 'Warranty')]

Products_in_Warranty = df_temp['Variable_A'].nunique()



df_temp = df_temp[df_temp['Date_1'].notnull()]
df_temp['Date_1'] = pd.to_datetime(df_temp['Date_1'])
df_temp['Quarter'] = df_temp['Date_1'].map(lambda x: int(pd.Timestamp(x).quarter) if ((x > datetime.datetime.now()) & (x < (datetime.datetime.now() + datetime.timedelta(days=180)))) else 0)


q = ceil(datetime.datetime.now().month / 3.0)
a = list(df_temp['Quarter'].unique())

if 0 in a:
    a.remove(0)
    
dict_warranty = {}
if q in a:
    y = str(datetime.datetime.now().year)
    name = "Products going out of warranty in Q" + str(q) + "'" + y[2:4]
    
    if df_temp[df_temp['Quarter'] == q].shape[0] > 0:
         dict_warranty[name] = df_temp[df_temp['Quarter'] == q]['Variable_A'].nunique()
    a.remove(q)


if len(a) > 0:
    for i in a:
        if i > q:
            yr = y[2:4]
            name = "Products going out of warranty in Q" + str(i) + "'" + y[2:4]
            dict_warranty[name] = df_temp[df_temp['Quarter'] == i]['Variable_A'].nunique()
        else:
            y = str(datetime.datetime.now().year + 1)
            name = "Products going out of warranty in Q" + str(i) + "'" + y[2:4]
            dict_warranty[name] = df_temp[df_temp['Quarter'] == i]['Variable_A'].nunique()





df_inwarranty = pd.Series(dict_warranty).to_frame()
df_inwarranty = df_inwarranty.reset_index(drop = False)
df_inwarranty.columns = ['Category', 'Count']
df_inwarranty = df_inwarranty.sort_values(by=['Count'], ascending=False)





    

# KPI - Contract status for Product out of warranty


Products_Out_Warranty = df[df['Variable_A'].notnull() & ((df['Status'] == 'Warranty') | (df['Status'] == 'Out of Warranty'))]['Variable_A'].nunique()

dict_outwarranty = {}

dict_outwarranty['In warranty'] = df[df['Status'] == 'Warranty']['Variable_A'].nunique()
dict_outwarranty['Out of Warranty'] = df[df['Status'] == 'Out of Warranty']['Variable_A'].nunique()


df_outwarranty = pd.Series(dict_outwarranty).to_frame()
df_outwarranty = df_outwarranty.reset_index(drop = False)
df_outwarranty.columns = ['Category', 'Count']
df_outwarranty = df_outwarranty.sort_values(by=['Count'], ascending=False)










# KPI - Age of products


df_temp = df[df['Date_2'].notnull()]
df_temp['Date_2'] = pd.to_datetime(df_temp['Date_2'])
df_temp['Cal_Age'] = (datetime.date.today() - df_temp['Date_2']).dt.days
df_temp['Cal_Age'] = df_temp['Cal_Age'].apply(lambda x: round(int(x) / 365))

def map_age(x):
    if x < 3:
        y = '< 2'
    elif (x > 2) & (x < 6):
        y = '3 - 5'
    elif (x > 5) & (x < 11):
        y = '6 - 10'
    elif (x > 10) & (x < 21):
        y = '11 - 20'
    elif x > 20:
        y = '> 20'
    else:
        y = 'Invalid'
    return y
        
        
df_temp['Cal_Age'] = df_temp['Cal_Age'].apply(lambda x: map_age(x))


# For all assets
df_age = df_temp.groupby(['Cal_Age'])['Variable_A'].nunique()
df_age = df_age.reset_index(drop = False)
df_age.columns = ['Age', 'Count of products']

if df[df['Date_2'].isnull()]['Variable_A'].count() > 0:
    df_age = df_age.append({'Age' : 'Blank', 'Count of Product' : df[df['Date_2'].isnull()]['Variable_A'].nunique()}, ignore_index=True)


Type_assets = dict(df_temp.groupby(['Cal_Age'])['Type'].nunique())

# Deliberately checked against "Variable_A" for Type count aginst Blank
if df[df['Date_2'].isnull()]['Variable_A'].count() > 0:
    Type_assets['Blank'] = df[df['Date_2'].isnull()]['Type'].nunique()

df_age['Type'] = df_age['Age'].map(Type_assets)



# For Out of Contract ################################



df_temp_new = df_temp[df_temp['Status'] == 'Out of contract']

df_OOC_age = df_temp_new.groupby(['Cal_Age'])['Variable_A'].nunique()
df_OOC_age = df_OOC_age.reset_index(drop = False)
df_OOC_age.columns = ['Age', 'Count of Product']

if df_temp_new[df_temp_new['Date_2'].isnull()]['Variable_A'].count() > 0:
    df_OOC_age = df_OOC_age.append({'Age' : 'Blank', 'Count of Product' : df_temp_new[df_temp_new['Date_2'].isnull()]['Variable_A'].nunique()}, ignore_index=True)


Type_OOC_assets = dict(df_temp_new.groupby(['Cal_Age'])['Type'].nunique())

if df_temp_new[df_temp_new['Date_2'].isnull()]['Variable_A'].count() > 0:
    Type_OOC_assets['Blank'] = df_temp_new[df_temp_new['Date_2'].isnull()]['Type'].nunique()

df_OOC_age['Type'] = df_OOC_age['Age'].map(Type_OOC_assets)




# For In Contract ################################


df_temp_new = df_temp[df_temp['Status'] == 'Contract']

df_IC_age = df_temp_new.groupby(['Cal_Age'])['Variable_A'].nunique()
df_IC_age = df_IC_age.reset_index(drop = False)
df_IC_age.columns = ['Age', 'Count of Product']

if df_temp_new[df_temp_new['Date_2'].isnull()]['Variable_A'].count() > 0:
    df_IC_age = df_IC_age.append({'Age' : 'Blank', 'Count of Product' : df_temp_new[df_temp_new['Date_2'].isnull()]['Variable_A'].nunique()}, ignore_index=True)


Type_IC_assets = dict(df_temp_new.groupby(['Cal_Age'])['Type'].nunique())

if df_temp_new[df_temp_new['Date_2'].isnull()]['Variable_A'].count() > 0:
    Type_IC_assets['Blank'] = df_temp_new[df_temp_new['Date_2'].isnull()]['Type'].nunique()

df_IC_age['Type'] = df_IC_age['Age'].map(Type_IC_assets)



# For In Warranty ################################


df_temp_new = df_temp[df_temp['Status'] == 'Warranty']

df_W_age = df_temp_new.groupby(['Cal_Age'])['Variable_A'].nunique()
df_W_age = df_W_age.reset_index(drop = False)
df_W_age.columns = ['Age', 'Count of Product']

if df_temp_new[df_temp_new['Date_2'].isnull()]['Variable_A'].count() > 0:
    df_W_age = df_W_age.append({'Age' : 'Blank', 'Count of Product' : df_temp_new[df_temp_new['Date_2'].isnull()]['Variable_A'].nunique()}, ignore_index=True)


Type_W_assets = dict(df_temp_new.groupby(['Cal_Age'])['Type'].nunique())

if df_temp_new[df_temp_new['Date_2'].isnull()]['Variable_A'].count() > 0:
    Type_W_assets['Blank'] = df_temp_new[df_temp_new['Date_2'].isnull()]['Type'].nunique()

df_W_age['Type'] = df_W_age['Age'].map(Type_W_assets)





#################################################################################
# Report generation
#################################################################################


out_file_name = str(df['Variable_D'].iloc[0].split()[0]) + '.xlsx'

writer = pd.ExcelWriter(out_file_name, engine='xlsxwriter')   
workbook=writer.book

workbook.formats[0].set_bg_color('white')

bold = workbook.add_format({'bold': True})

italic_small = workbook.add_format({'font_name': 'Calibri', 'font_size' : 9, 'italic' : True})
italic_large = workbook.add_format({'font_name': 'Calibri', 'font_size' : 11, 'italic' : True})


underline = workbook.add_format({'bold': True, 'underline' : 'single'})


worksheet=workbook.add_worksheet('Dashboard')
writer.sheets['Dashboard'] = worksheet
#

worksheet.set_column('A:A', width = 9)
worksheet.set_column('B:B', width = 32)
worksheet.set_column('C:C', width = 16)
worksheet.set_column('D:D', width = 17)
worksheet.set_column('E:E', width = 1)
worksheet.set_column('F:F', width = 1)
worksheet.set_column('G:G', width = 26)
worksheet.set_column('H:H', width = 13)
worksheet.set_column('I:I', width = 16)
worksheet.set_column('J:J', width = 1)


row_nbr = 5
df_sys.to_excel(writer,sheet_name='Dashboard', startrow = row_nbr, startcol = 1, index=False)   
cache_sys_row_start = 7
cache_sys_row_end = cache_sys_row_start + df_sys.shape[0] - 1




row_nbr = row_nbr + 3 + df_sys.shape[0]
df_inwarranty.to_excel(writer, sheet_name='Dashboard', startrow = row_nbr, startcol = 1, index=False, header=False)
cache_inwarranty_row_start = cache_sys_row_end + 2
cache_inwarranty_row_end = cache_inwarranty_row_start + df_inwarranty.shape[0]




row_nbr = row_nbr + 2 + df_inwarranty.shape[0]
df_outwarranty.to_excel(writer, sheet_name='Dashboard', startrow = row_nbr, startcol = 1, index=False, header=False)
cache_outwarranty_row_start = cache_inwarranty_row_end + 2
cache_outwarranty_row_end = cache_outwarranty_row_start + df_outwarranty.shape[0]




row_nbr = row_nbr + 3 + df_outwarranty.shape[0]
df_age.to_excel(writer, sheet_name='Dashboard', startrow = row_nbr, startcol = 1, index=False, header=False)
cache_age_row_start = cache_outwarranty_row_end + 4
cache_age_row_end = cache_age_row_start + df_age.shape[0] - 1



row_nbr = row_nbr + 3 + df_age.shape[0]
df_OOC_age.to_excel(writer, sheet_name='Dashboard', startrow = row_nbr, startcol = 1, index=False, header=False)
cache_OOC_age_row_start = cache_age_row_end + 4
cache_OOC_age_row_end = cache_OOC_age_row_start + df_OOC_age.shape[0] - 1



row_nbr = row_nbr + 3 + df_OOC_age.shape[0]
df_IC_age.to_excel(writer, sheet_name='Dashboard', startrow = row_nbr, startcol = 1, index=False, header=False)
cache_IC_age_row_start = cache_OOC_age_row_end + 4
cache_IC_age_row_end = cache_IC_age_row_start + df_IC_age.shape[0] - 1




row_nbr = row_nbr + 3 + df_IC_age.shape[0]
df_W_age.to_excel(writer, sheet_name='Dashboard', startrow = row_nbr, startcol = 1, index=False, header=False)
cache_W_age_row_start = cache_IC_age_row_end + 4
cache_W_age_row_end = cache_W_age_row_start + df_W_age.shape[0] - 1


worksheet.write(0, 1, 'Customer:', bold)
worksheet.write(0, 2, str(df['Variable_D'].iloc[0].split()[0]), bold)
worksheet.write(2, 1, 'Total Product Analysis', underline)
worksheet.write(4, 1, 'Unique Products', bold)
worksheet.write(4, 2, unique_asset_count)


worksheet.write(cache_inwarranty_row_start-1, 1, 'Products in Warranty', bold)
worksheet.write(cache_inwarranty_row_start-1, 2, Products_in_Warranty)



worksheet.write(cache_outwarranty_row_start-1, 1, 'Products out of warranty', bold)
worksheet.write(cache_outwarranty_row_start-1, 2, Products_Out_Warranty)





worksheet.write(cache_age_row_start-2, 1, 'Age of Products', bold)
worksheet.write(cache_age_row_start-2, 2, 'Count of Product', bold)
worksheet.write(cache_age_row_start-2, 3, 'Unique # Type', bold)

worksheet.write(cache_OOC_age_row_start-3, 1, 'Out of Contract')
worksheet.write(cache_OOC_age_row_start-2, 1, 'Age of Products', bold)
worksheet.write(cache_OOC_age_row_start-2, 2, 'Count of Product', bold)
worksheet.write(cache_OOC_age_row_start-2, 3, 'Unique # Type', bold)




worksheet.write(cache_IC_age_row_start-3, 1, 'Contract')
worksheet.write(cache_IC_age_row_start-2, 1, 'Age of Products', bold)
worksheet.write(cache_IC_age_row_start-2, 2, 'Count of Product', bold)
worksheet.write(cache_IC_age_row_start-2, 3, 'Unique # Type', bold)



worksheet.write(cache_W_age_row_start-3, 1, 'Warranty')
worksheet.write(cache_W_age_row_start-2, 1, 'Age of Products', bold)
worksheet.write(cache_W_age_row_start-2, 2, 'Count of Product', bold)
worksheet.write(cache_W_age_row_start-2, 3, 'Unique # Type', bold)



                
writer.save()


# Set the borders


from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl import load_workbook
wb = load_workbook(filename=out_file_name)
ws = wb.worksheets[0]


def set_border(ws, cell_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws.iter_rows(cell_range)
    for row in rows:
        for cell in row:
            cell.border = border



cell_set = 'B' + str(cache_sys_row_start) + ':D' + str(cache_sys_row_end)
set_border(ws, cell_set)

cell_set = 'B' + str(cache_inwarranty_row_start) + ':C' + str(cache_inwarranty_row_end)
set_border(ws, cell_set)
for x in range(cache_inwarranty_row_start+1, cache_inwarranty_row_end+1):
    cell = 'B' + str(x)
    ws[cell].font = Font(name='Calibri', size=9, italic=True)
    ws[cell].alignment = Alignment(horizontal='right')
    
    
    

cell_set = 'B' + str(cache_outwarranty_row_start) + ':C' + str(cache_outwarranty_row_end)
set_border(ws, cell_set)
for x in range(cache_outwarranty_row_start+1, cache_outwarranty_row_end+1):
    cell = 'B' + str(x)
    ws[cell].font = Font(name='Calibri', size=9, italic=True)
    ws[cell].alignment = Alignment(horizontal='right')



cell_set = 'B' + str(cache_age_row_start) + ':D' + str(cache_age_row_end)
set_border(ws, cell_set)
for x in range(cache_age_row_start+1, cache_age_row_end+1):
    cell = 'B' + str(x)
    ws[cell].font = Font(name='Calibri', size=11, italic=True)
    ws[cell].alignment = Alignment(horizontal='left')




cell_set = 'B' + str(cache_OOC_age_row_start) + ':D' + str(cache_OOC_age_row_end)
set_border(ws, cell_set)
for x in range(cache_OOC_age_row_start+1, cache_OOC_age_row_end+1):
    cell = 'B' + str(x)
    ws[cell].font = Font(name='Calibri', size=11, italic=True)
    ws[cell].alignment = Alignment(horizontal='left')



cell_set = 'B' + str(cache_IC_age_row_start) + ':D' + str(cache_IC_age_row_end)
set_border(ws, cell_set)
for x in range(cache_IC_age_row_start+1, cache_IC_age_row_end+1):
    cell = 'B' + str(x)
    ws[cell].font = Font(name='Calibri', size=11, italic=True)
    ws[cell].alignment = Alignment(horizontal='left')



cell_set = 'B' + str(cache_W_age_row_start) + ':D' + str(cache_W_age_row_end)
set_border(ws, cell_set)
for x in range(cache_W_age_row_start+1, cache_W_age_row_end+1):
    cell = 'B' + str(x)
    ws[cell].font = Font(name='Calibri', size=11, italic=True)
    ws[cell].alignment = Alignment(horizontal='left')



def set_bg_color(ws, cell_range):
    rows = ws.iter_rows(cell_range)
    for row in rows:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor='FFFFFF')


set_bg_color(ws, 'A1:Z200')










wb.save(out_file_name)




